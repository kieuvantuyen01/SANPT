import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
import sys
# from pysat.formula import CNF
from pysat.solvers import Glucose3, Solver
# import time
from threading import Timer
import os
import ast

sat_solver = Glucose3
time_budget = 600  # Set your desired time budget in seconds
type = "es3"
id_counter = 1

# Open the log file in append mode
log_file = open('console.log', 'a')

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")


# Define a custom print function that writes to both console and log file
def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file = log_file, **kwargs)
    log_file.flush()

def check_overlap(task1, task2):
    # Suppose: task1 = (r1, e1, d1), task2 = (r2, e2, d2)
    # r1_min = r1, r1_max = d1 - e1, d1_min = r1 + e1, d1_max = d1
    # r2_min = r2, r2_max = d2 - e2, d2_min = r2 + e2, d2_max = d2
    # task1 and task2 are overlapped if: 
    # 1. d2_min >= r1_max and r2_max <= d1_min
    # 2. d1_min >= r2_max and r1_max <= d2_min
    # => r2 + e2 >= d1 - e1 and d2 - e2 <= r1 + e1 or r1 + e1 >= d2 - e2 and d1 - e1 <= r2 + e2
    if task2[0] + task2[1] > task1[2] - task1[1] and task2[2] - task2[1] < task1[0] + task1[1]:
        return True
    if task1[0] + task1[1] > task2[2] - task2[1] and task1[2] - task1[1] < task2[0] + task2[1]:
        return True
    return False

def encode_problem_es3(tasks, resources):
    max_time = max(task[2] for task in tasks)

    # Variables u[i][j] for task i accessing resource j
    u = [[i * resources + j + 1 for j in range(resources)] for i in range(len(tasks))]

    # Variables z[i][t] for task i accessing some resource at time t
    z = [[len(tasks) * resources + i * tasks[i][2] + t + 1 for t in range(tasks[i][2])] for i in range(len(tasks))]

    # Variables D[i][j][t] for non-preemptive access of resource j by task i starting at time t
    D = [[[len(tasks) * (resources + max_time) + i * resources * max_time + j * max_time + t + 1 
          for t in range(max_time)] for j in range(resources)] for i in range(len(tasks))]

#    # Overlapping: check each pair of tasks to see if they are overlap time, u_i1j -> -u_i2j
#     for i in range(len(tasks)):
#         for ip in range(i + 1, len(tasks)):
#             if check_overlap(tasks[i], tasks[ip]):
#                 for j in range(resources):
#                     sat_solver.add_clause([-u[i][j], -u[ip][j]])
#                     # print(f"Added clause D0: -u{i+1}{j+1} -u{ip+1}{j+1}")

#     # Symmetry breaking 1: Assign the tasks to resources if have r_max <= d_min (min of all tasks)
#     d_min = min(task[2] for task in tasks)
#     fixed_tasks = []
#     for i in range(len(tasks)):
#         if tasks[i][2] - tasks[i][1] <= d_min:
#             fixed_tasks.append(i)
#     # Assign each task in fixed_tasks to a resource
#     for j, i in enumerate(fixed_tasks):
#         if j < resources:
#             sat_solver.add_clause([u[i][j]])
#         # print(f"Added clause S1: u{i+1}{j+1}")
    
#     # Symmetry breaking 2: if each task i has t in range(r_max, d_min), then z[i][t] = True
#     # for j in range(resources):
#     for i in range(len(tasks)):
#         for t in range(tasks[i][2] - tasks[i][1], tasks[i][0] + tasks[i][1]):
#             sat_solver.add_clause([z[i][t]])
#             # print(f"Added clause S2: -u{i+1}{j+1}, z{i+1}{t}")

    # D1: Task i should not access two resources at the same time
    for i in range(len(tasks)):
        for j in range(resources):
            for jp in range(j + 1, resources):
                sat_solver.add_clause([-u[i][j], -u[i][jp]])
                # print(f"Added clause D1: -u{i+1}{j+1} -u{i+1}{jp+1}")

    # D2: Each task must get some resource
    for i in range(len(tasks)):
        # sat_solver.add_clause([u[i][j] for j in range(resources)])
        # print(f"Added clause: u{i}0 u{i}1")
        clause = []
        clause_str = []
        for j in range(resources):
            clause.append(u[i][j])
            clause_str.append(f"u{i+1}{j+1}")
        sat_solver.add_clause(clause)
        # print(f"Added clause D2: {clause_str}")

    # D3: A resource can only be held by one task at a time
    for i in range(len(tasks)):
        for ip in range(i + 1, len(tasks)):
            for j in range(resources):
                for t in range(tasks[i][0], min(tasks[i][2], tasks[ip][2])):
                    sat_solver.add_clause([-z[i][t], -u[i][j], -z[ip][t], -u[ip][j]])
                    # print(f"Added clause D3: -z{i+1}{t} -u{i+1}{j+1} -z{ip+1}{t} -u{ip+1}{j+1}")

    # D4: Each task must have exactly one start time for accessing a resource non-preemptively
    for i in range(len(tasks)):
        # sat_solver.add_clause([D[i][j][t] for j in range(resources) for t in range(tasks[i][2] - tasks[i][1] + 1)])
        # print(f"Added clause D4: D{i}{j}{t}")
        clause = []
        clause_str = []
        for j in range(resources):
            for t in range(tasks[i][2] - tasks[i][1] + 1):
                clause.append(D[i][j][t])
                clause_str.append(f"D{i+1}{j+1}{t}")
        sat_solver.add_clause(clause)
        # print(f"Added clause D4: {clause_str}")

    # D5: Linking start variable to z and u variables
    for i in range(len(tasks)):
        for j in range(resources):
            for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1):
                # Reverse implication
                clause = [D[i][j][t]]
                clause.extend([-u[i][j]])
                clause_str = []
                clause_str.append(f"D{i+1}{j+1}{t}")
                clause_str.append(f"-u{i+1}{j+1}")

                # If D[i][j][t] is true, the task must hold the resource for its entire duration
                for tp in range(t, t + tasks[i][1]):
                    if tp < max_time:
                        sat_solver.add_clause([-D[i][j][t], z[i][tp]])
                        clause.append(-z[i][tp])
                        clause_str.append(f"-z{i+1}{tp}")
                        # print(f"Added clause D5: -D{i+1}{j+1}{t} z{i+1}{tp}")
                sat_solver.add_clause([-D[i][j][t], u[i][j]])
                # clause.append(-u[i][j])
                # clause_str.append(f"-u{i+1}{j+1}")
                # print(f"Added clause D5: -D{i+1}{j+1}{t} u{i+1}{j+1}")

                # If D[i][j][t] is true, the task must not hold the resource before t
                for tp in range(tasks[i][0], t):
                    sat_solver.add_clause([-D[i][j][t], -z[i][tp]])
                    clause.append(z[i][tp])
                    clause_str.append(f"z{i+1}{tp}")
                    # print(f"Added clause D5: -D{i+1}{j+1}{t} -z{i+1}{tp}")

                # If D[i][j][t] is true, the task must not hold the resource after t + e_i - 1
                for tp in range(t + tasks[i][1], tasks[i][2]):
                    if tp < max_time:
                        sat_solver.add_clause([-D[i][j][t], -z[i][tp]])
                        clause.append(z[i][tp])
                        clause_str.append(f"z{i+1}{tp}")
                        # print(f"Added clause D5: -D{i+1}{j+1}{t} -z{i+1}{tp}")

                sat_solver.add_clause(clause)
                # print(f"Added clause D5: {clause_str}")

    return u, z, D

def interrupt(solver):
    solver.interrupt()

def solve_es3(tasks, resources):
    global sat_solver

    # with Solver(name="glucose4") as solver:
    sat_solver = Glucose3(use_timer=True)
    
    u, z, D = encode_problem_es3(tasks, resources)

    # start_time = time.time()
    num_variables = sat_solver.nof_vars()
    num_clauses = sat_solver.nof_clauses()

    print_to_console_and_log(f"Num of variables: {num_variables}")
    print_to_console_and_log(f"Num of clauses: {num_clauses}")

    timer = Timer(time_budget, interrupt, [sat_solver])
    timer.start()
    result = sat_solver.solve_limited(expect_interrupt = True)

    solve_time = float(format(sat_solver.time(), ".6f"))

    res = ""
    if result:
        model = sat_solver.get_model()
        if model is None:
            print("Time out")
            res = "Time out"
        else:
            print("SAT")
            res = "SAT"
            for i in range(len(tasks)):
                for j in range(resources):
                    if model[u[i][j] - 1] > 0:
                        print_to_console_and_log(f"Task {i} is assigned to resource {j}")
                for t in range(tasks[i][0], tasks[i][2]):
                    if model[z[i][t] - 1] > 0:
                        print_to_console_and_log(f"Task {i} is accessing a resource at time {t}")
                for j in range(resources):
                    for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1):
                        if model[D[i][j][t] - 1] > 0:
                            print_to_console_and_log(f"Task {i} starts non-preemptive access of resource {j} at time {t}")
    else:
        print("UNSAT")
        res = "UNSAT"

    timer.cancel()
    sat_solver.delete()

    return res, solve_time, num_variables, num_clauses
    
def process_input_files(input_folder, resources=200):
    global id_counter, type

    # results = {}
    for filename in os.listdir(input_folder):
        if filename.endswith(".txt"):
            file_path = os.path.join(input_folder, filename)
            with open(file_path, 'r') as f:
                num_tasks = int(f.readline().strip())
                tasks = ast.literal_eval(f.readline().strip())
                print(f"tasks: {tasks}")

            print(f"Processing {filename}...")
            # res, solve_time, num_variables, num_clauses = solve_es3(tasks, num_tasks)
            res, solve_time, num_variables, num_clauses = solve_es3(tasks, resources)
            # results[filename] = {
            #     "result": res,
            #     "time": float(solve_time),
            #     "num_variables": num_variables,
            #     "num_clauses": num_clauses
            # }
            result_dict = {
                "ID": id_counter,
                "Problem": os.path.basename(filename),
                "Type": type,
                "Time": solve_time,
                "Result": res,
                "Variables": num_variables,
                "Clauses": num_clauses
            }
            write_to_xlsx(result_dict)
            id_counter += 1

    # return results

# Main execution
input_folder = "input/" + sys.argv[1]
process_input_files(input_folder)

log_file.close()

# export results to excel

# # Print summary of results
# print("\nSummary of results:")
# # Print sum time of satisfiable results, sum number of variables, sum number of clauses of all problems
# satisfiable_results = [result for result in results.values() if result["result"] == "SAT"]
# sum_of_time = sum(result["time"] for result in satisfiable_results)
# sum_of_variables = sum(result["num_variables"] for result in results.values())
# sum_of_clauses = sum(result["num_clauses"] for result in results.values())
# print(f"Average time of satisfiable results: {sum_of_time:.6f} s")
# print(f"Average number of variables: {sum_of_variables}")
# print(f"Average number of clauses: {sum_of_clauses}")